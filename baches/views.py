from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.http import HttpResponseForbidden, HttpResponse
from django.contrib.auth import logout, login
from django.core.serializers.json import DjangoJSONEncoder
from django.utils import timezone
from django.db.models import Q, Count
from django.urls import reverse
from django.contrib.auth.views import LoginView
from django.contrib import messages
from django.core.mail import send_mail
from django.contrib.auth.models import User
from django.conf import settings
from django.contrib.auth.hashers import make_password, check_password
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from reportlab.lib.pagesizes import A4, landscape
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle
from reportlab.lib import colors

import json
import csv
import random
import secrets
from datetime import timedelta

# Importaciones Locales de la Aplicación
from .models import (
    Bache,
    HistorialBache,
    Notificacion,
    Perfil,
    CodigoEmail,
    CodigoVerificacionEmail,
    PasswordResetCode,
    ImagenBache
)
from .forms import BacheForm, RegistroForm, PasswordResetRequestForm, PasswordResetConfirmForm
from .decorators import solo_municipio

#CARGAR DATOS EN PRINCIPAL
def lista_baches(request):
    # Filtros desde la querystring (?estado=...&severidad=...&mios=1)
    estado = request.GET.get("estado", "")
    severidad = request.GET.get("severidad", "")
    mios = request.GET.get("mios", "")

    # Query
    baches_qs = Bache.objects.all().order_by("-fecha_creacion")

    # Aplicamos filtros si vinieron
    if estado:
        baches_qs = baches_qs.filter(estado=estado)

    if severidad:
        baches_qs = baches_qs.filter(severidad=severidad)
    
    if mios and request.user.is_authenticated:
        baches_qs = baches_qs.filter(vecino=request.user)

    # Para la tabla usamos el queryset filtrado
    baches = baches_qs
    # Para el mapa, del mismo queryset filtrado tomamos solo los que tienen coords
    baches_con_coords = (
        baches_qs
        .exclude(latitud__isnull=True)
        .exclude(longitud__isnull=True)
    )
    # Pasamos datos simples a JS
    baches_data = [
        {
            "id": b.id,
            "titulo": b.titulo,
            "latitud": b.latitud,
            "longitud": b.longitud,
            "severidad": b.severidad,
            "estado": b.estado,
            "imagen_url": b.imagen.url if b.imagen else None,
        }
        for b in baches_con_coords
    ]
    # Enviamos también los filtros actuales para mantenerlos seleccionados en la UI
    return render(
        request,
        "baches/lista_baches.html",
        {
            "baches": baches,
            "baches_data": baches_data,
            "estado_actual": estado,
            "severidad_actual": severidad,
            "mios_actual": mios,
        }
    )


# CREAR
@login_required
def crear_bache(request):
    if request.method == "POST":
        form = BacheForm(request.POST, request.FILES)
        if form.is_valid():
            bache = form.save(commit=False)
            bache.vecino = request.user
            bache.save()

            # guardar múltiples imágenes
            for f in request.FILES.getlist("imagenes"):
                ImagenBache.objects.create(bache=bache, imagen=f)

            return redirect("lista_baches")
    else:
        form = BacheForm()

    return render(request, "baches/crear_bache.html", {"form": form})



# MUNICIPIO
@login_required
def panel_municipio(request):
    if not hasattr(request.user, "perfil") or request.user.perfil.rol != "municipio":
        return HttpResponseForbidden("No tenés permisos para ver esta página.")

    # =========================
    # POST: actualizar en masa
    # =========================
    
    por_barrio_labels = []
    por_barrio_values = []
    
    if request.method == "POST":
        for key, value in request.POST.items():
            if not key.startswith("estado_"):
                continue

            bache_id = key.split("_")[1]
            nuevo_estado = (value or "").strip()

            comentario_key = f"comentario_{bache_id}"
            nuevo_comentario = (request.POST.get(comentario_key, "") or "").strip()

            try:
                bache = Bache.objects.get(id=bache_id)
            except Bache.DoesNotExist:
                continue

            estado_anterior = bache.estado
            comentario_anterior = getattr(bache, "comentario_municipio", "") or ""

            hubo_cambio = False

            # cambio de estado
            if nuevo_estado and nuevo_estado != estado_anterior:
                bache.estado = nuevo_estado
                hubo_cambio = True

                if hasattr(bache, "fecha_resolucion"):
                    if nuevo_estado == "resuelto":
                        bache.fecha_resolucion = timezone.now()
                    else:
                        bache.fecha_resolucion = None

            # cambio de comentario
            if hasattr(bache, "comentario_municipio") and nuevo_comentario != comentario_anterior:
                bache.comentario_municipio = nuevo_comentario
                hubo_cambio = True

            if not hubo_cambio:
                continue

            bache.save()

            # Historial
            HistorialBache.objects.create(
                bache=bache,
                actor=request.user,
                estado_anterior=estado_anterior,
                estado_nuevo=bache.estado,
                comentario_anterior=comentario_anterior,
                comentario_nuevo=getattr(bache, "comentario_municipio", "") or "",
            )

            # Notificación al vecino (sistema + mail si ya lo tenés implementado)
            if bache.vecino:
                cambio_estado = (bache.estado != estado_anterior)
                cambio_coment = (
                    hasattr(bache, "comentario_municipio")
                    and nuevo_comentario != comentario_anterior
                    and nuevo_comentario != ""
                )

                if cambio_estado and cambio_coment:
                    tipo = "mixto"
                    mensaje = (
                        f"Tu reclamo '{bache.titulo}' cambió a {bache.get_estado_display()}.\n"
                        f"Comentario del municipio: {nuevo_comentario}"
                    )
                elif cambio_estado:
                    tipo = "estado"
                    mensaje = (
                        f"Tu reclamo '{bache.titulo}' cambió a {bache.get_estado_display()}.\n"
                        f"Comentario: {nuevo_comentario}" if nuevo_comentario else
                        f"Tu reclamo '{bache.titulo}' cambió a {bache.get_estado_display()}."
                    )
                elif cambio_coment:
                    tipo = "comentario"
                    mensaje = (
                        f"El municipio dejó un comentario en tu reclamo '{bache.titulo}':\n"
                        f"{nuevo_comentario}"
                    )
                else:
                    tipo = "info"
                    mensaje = f"Hubo una actualización en tu reclamo '{bache.titulo}'."

                Notificacion.objects.create(
                    destinatario=bache.vecino,
                    bache=bache,
                    tipo=tipo,
                    mensaje=mensaje
                )

                enviar_mail_cambio_bache(
                destinatario=bache.vecino,
                asunto="Actualización de tu reclamo - PozosYa",
                cuerpo=mensaje
                )

        # IMPORTANTE: volver al panel conservando filtros GET
        qs = request.META.get("QUERY_STRING", "")

    # =========================
    # GET: filtros + orden
    # =========================
    estado = request.GET.get("estado", "").strip()
    severidad = request.GET.get("severidad", "").strip()
    barrio = request.GET.get("barrio", "").strip()
    q = request.GET.get("q", "").strip()
    orden = request.GET.get("orden", "").strip()  # nuevo: orden por votos / fecha

    qs = Bache.objects.all()

    # annotate votos (si tu upvotes es ManyToMany/related_name="upvotes")
    qs = qs.annotate(votos_count=Count("upvotes", distinct=True))

    if estado:
        qs = qs.filter(estado=estado)
    if severidad:
        qs = qs.filter(severidad=severidad)
    if barrio:
        qs = qs.filter(barrio=barrio)
    if q:
        qs = qs.filter(
            Q(titulo__icontains=q) |
            Q(calle__icontains=q) |
            Q(altura__icontains=q)
        )

    # Orden
    if orden == "mas_votados":
        qs = qs.order_by("-votos_count", "-fecha_creacion")
    elif orden == "menos_votados":
        qs = qs.order_by("votos_count", "-fecha_creacion")
    else:
        # default: más nuevos
        qs = qs.order_by("-fecha_creacion")

    baches = qs

            # Top barrios (para que no explote el gráfico si hay 200 barrios)
    por_barrio = (
        baches
        .exclude(barrio__isnull=True)
        .exclude(barrio__exact="")
        .values("barrio")
        .annotate(c=Count("id"))
        .order_by("-c")
    )

    top_n = 12
    top = list(por_barrio[:top_n])
    resto = list(por_barrio[top_n:])
    otros_total = sum(x["c"] for x in resto)

    por_barrio_chart = top + ([{"barrio": "Otros", "c": otros_total}] if otros_total else [])

    por_barrio_labels = [x["barrio"] for x in por_barrio_chart]
    por_barrio_values = [x["c"] for x in por_barrio_chart]


    # Barrios para el select (solo los existentes)
    barrios = (
        Bache.objects.exclude(barrio="")
        .values_list("barrio", flat=True)
        .distinct()
        .order_by("barrio")
    )

    # Dashboard (contadores del total sin filtros o con filtros? acá con filtros)
    total = baches.count()
    nuevos = baches.filter(estado="nuevo").count()
    en_gestion = baches.filter(estado="en_gestion").count()
    resueltos = baches.filter(estado="resuelto").count()

    return render(request, "baches/panel_municipio.html", {
        "baches": baches,
        "ESTADO_CHOICES": Bache.ESTADO_CHOICES,
        "SEVERIDAD_CHOICES": Bache.SEVERIDAD_CHOICES,
        "barrios": barrios,
        "estado_actual": estado,
        "severidad_actual": severidad,
        "barrio_actual": barrio,
        "q_actual": q,
        "orden_actual": orden,
        "total": total,
        "nuevos": nuevos,
        "en_gestion": en_gestion,
        "resueltos": resueltos,
        "por_barrio_labels": por_barrio_labels,
        "por_barrio_values": por_barrio_values,
    })




# LOGOUT
def logout_view(request):
    logout(request)  # borra la sesión del usuario
    return redirect('lista_baches')  # siempre vuelve al inicio



# REGISTRARSE
def registro(request):
    if request.method == "POST":
        form = RegistroForm(request.POST)
        if form.is_valid():
            # 1) Crear usuario pero INACTIVO
            user = form.save(commit=False)
            user.email = form.cleaned_data.get("email", "")
            user.first_name = form.cleaned_data.get("first_name", "")
            user.last_name = form.cleaned_data.get("last_name", "")
            user.is_active = False
            user.save()

            # 2) Crear perfil vecino
            Perfil.objects.create(
                user=user,
                rol="vecino",
                dni=form.cleaned_data.get("dni", ""),
                fecha_nacimiento=form.cleaned_data.get("fecha_nacimiento"),
                telefono=form.cleaned_data.get("telefono", ""),
                domicilio=form.cleaned_data.get("domicilio", ""),
            )

            # 3) Generar código y guardarlo
            codigo = f"{random.randint(0, 999999):06d}"
            expira = timezone.now() + timedelta(minutes=15)

            CodigoEmail.objects.create(
                user=user,
                tipo="verify",
                codigo=codigo,
                expira_en=expira,
                usado=False,
            )

            # 4) Enviar mail real
            send_mail(
                subject="Código de verificación - PozosYa",
                message=f"Tu código es: {codigo}\nVence en 15 minutos.",
                from_email=settings.DEFAULT_FROM_EMAIL,
                recipient_list=[user.email],
                fail_silently=False,
            )

            # 5) Guardar en sesión para simplificar el flujo
            request.session["verify_user_id"] = user.id

            messages.success(request, "Te enviamos un código al email. Ingresalo para activar tu cuenta.")
            return redirect("verificar_email")
        else:
            messages.error(request, "Revisá los datos: hay errores en el formulario.")
    else:
        form = RegistroForm()

    return render(request, "registration/registro.html", {"form": form})



# EXPORTAR A CSV
@login_required
def exportar_baches_csv(request):
    if not hasattr(request.user, "perfil") or request.user.perfil.rol != "municipio":
        return HttpResponseForbidden("No tenés permisos.")

    baches = _get_baches_filtrados_para_export(request)

    response = HttpResponse(content_type="text/csv")
    response["Content-Disposition"] = 'attachment; filename="baches_pozosya.csv"'

    writer = csv.writer(response)
    writer.writerow([
        "ID", "Titulo", "Calle", "Altura", "Barrio", "Severidad",
        "Estado", "Votos", "Latitud", "Longitud", "Vecino", "Fecha"
    ])

    for b in baches:
        writer.writerow([
            b.id, b.titulo, b.calle, b.altura, b.barrio,
            b.get_severidad_display(), b.get_estado_display(),
            getattr(b, "votos_count", 0),
            b.latitud, b.longitud,
            b.vecino.username if b.vecino else "",
            b.fecha_creacion.strftime("%d/%m/%Y %H:%M"),
        ])

    return response



# EXPORTAR EXCEL´S
@login_required
def exportar_baches_excel(request):
    if not hasattr(request.user, "perfil") or request.user.perfil.rol != "municipio":
        return HttpResponseForbidden("No tenés permisos.")

    baches = _get_baches_filtrados_para_export(request)

    wb = Workbook()
    ws = wb.active
    ws.title = "Baches"

    headers = [
        "ID", "Titulo", "Calle", "Altura", "Barrio", "Severidad",
        "Estado", "Votos", "Latitud", "Longitud", "Vecino", "Fecha"
    ]
    ws.append(headers)

    for b in baches:
        ws.append([
            b.id,
            b.titulo,
            b.calle,
            b.altura,
            b.barrio,
            b.get_severidad_display(),
            b.get_estado_display(),
            int(getattr(b, "votos_count", 0)),
            b.latitud,
            b.longitud,
            b.vecino.username if b.vecino else "",
            b.fecha_creacion.strftime("%d/%m/%Y %H:%M"),
        ])

    ws.freeze_panes = "A2"

    # Ajuste simple de ancho de columnas
    for col_idx, col_name in enumerate(headers, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = max(12, len(col_name) + 2)

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename="baches_pozosya.xlsx"'
    wb.save(response)
    return response



# EXPORTAR PDS´S
@login_required
def exportar_baches_pdf(request):
    if not hasattr(request.user, "perfil") or request.user.perfil.rol != "municipio":
        return HttpResponseForbidden("No tenés permisos.")

    baches = _get_baches_filtrados_para_export(request)

    response = HttpResponse(content_type="application/pdf")
    response["Content-Disposition"] = 'attachment; filename="baches_pozosya.pdf"'

    doc = SimpleDocTemplate(response, pagesize=landscape(A4), rightMargin=20, leftMargin=20, topMargin=20, bottomMargin=20)

    data = [[
        "ID", "Titulo", "Ubicacion", "Barrio", "Severidad", "Estado", "Votos", "Fecha"
    ]]

    for b in baches:
        ubicacion = f"{b.calle} {b.altura}".strip()
        data.append([
            str(b.id),
            (b.titulo or "")[:60],
            ubicacion[:45],
            (b.barrio or "")[:25],
            b.get_severidad_display(),
            b.get_estado_display(),
            str(getattr(b, "votos_count", 0)),
            b.fecha_creacion.strftime("%d/%m/%Y %H:%M"),
        ])

    table = Table(data, repeatRows=1)
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1f6f9b")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 9),
        ("GRID", (0, 0), (-1, -1), 0.25, colors.lightgrey),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.whitesmoke, colors.white]),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
    ]))

    doc.build([table])
    return response



# FILTROS PARA CUANDO EXPORTE
def _get_baches_filtrados_para_export(request):
    estado = request.GET.get("estado", "").strip()
    severidad = request.GET.get("severidad", "").strip()
    barrio = request.GET.get("barrio", "").strip()
    q = request.GET.get("q", "").strip()
    orden = request.GET.get("orden", "").strip()

    qs = Bache.objects.all().annotate(votos_count=Count("upvotes", distinct=True))

    if estado:
        qs = qs.filter(estado=estado)
    if severidad:
        qs = qs.filter(severidad=severidad)
    if barrio:
        qs = qs.filter(barrio=barrio)
    if q:
        qs = qs.filter(
            Q(titulo__icontains=q) |
            Q(calle__icontains=q) |
            Q(altura__icontains=q)
        )

    if orden == "mas_votados":
        qs = qs.order_by("-votos_count", "-fecha_creacion")
    elif orden == "menos_votados":
        qs = qs.order_by("votos_count", "-fecha_creacion")
    else:
        qs = qs.order_by("-fecha_creacion")

    return qs





# DETALLE
def detalle_bache(request, pk):
    bache = get_object_or_404(Bache, pk=pk)

    historial = (
        bache.historial
        .select_related("actor")
        .order_by("-fecha")
    )

    return render(
        request,
        "baches/detalle_bache.html",
        {"bache": bache, "historial": historial}
    )



# NOTIFICACIONES
@login_required
def mis_notificaciones(request):
    # Trae todas las notificaciones del usuario logueado
    notifs = (
        Notificacion.objects.filter(destinatario=request.user).order_by("-fecha")
    )

    # Cantidad no leídas para mostrar en la página también
    no_leidas = notifs.filter(leida=False).count()

    return render(request, "baches/notificaciones.html", {"notificaciones": notifs, "no_leidas": no_leidas})



# MARCAR NOTIFICACIONES COMO LEÍDAS
@login_required
def marcar_notificacion_leida(request, notif_id):
    try:
        notif = Notificacion.objects.get(id=notif_id, destinatario=request.user)
        notif.leida = True
        notif.save()
    except Notificacion.DoesNotExist:
        pass

    # Si querés que al tocarla te lleve al detalle del bache:
    return redirect("detalle_bache", pk=notif.bache.id)



# A DEFINIR
@login_required
def mis_baches(request):
    baches = Bache.objects.filter(vecino=request.user).order_by("-fecha_creacion")

    baches_con_coords = (
        baches
        .exclude(latitud__isnull=True)
        .exclude(longitud__isnull=True)
    )

    baches_data = [
        {
            "id": b.id,
            "titulo": b.titulo,
            "latitud": b.latitud,
            "longitud": b.longitud,
            "severidad": b.severidad,
            "estado": b.estado,
            "imagen_url": b.imagen.url if b.imagen else None,
        }
        for b in baches_con_coords
    ]

    return render(request, "baches/mis_baches.html", {
        "baches": baches,
        "baches_data": baches_data,
    })



# UPVOTE
@login_required
def toggle_upvote(request, pk):
    bache = get_object_or_404(Bache, pk=pk)

    if request.user in bache.upvotes.all():
        bache.upvotes.remove(request.user)
    else:
        bache.upvotes.add(request.user)

    return redirect("detalle_bache", pk=pk)



# MENSAJE DE ERROR
class CustomLoginView(LoginView):
    template_name = "registration/login.html"

    def form_invalid(self, form):
        messages.error(self.request, "Usuario o contraseña incorrectos.")
        return super().form_invalid(form)



# VERIFICAR EMAIL
def verificar_email(request):
    user_id = request.session.get("verify_user_id")
    if not user_id:
        messages.error(request, "No hay verificación pendiente.")
        return redirect("login")

    user = get_object_or_404(User, id=user_id)

    if request.method == "POST":
        codigo_ingresado = request.POST.get("codigo", "").strip()
        cod = (CodigoEmail.objects
            .filter(user=user, tipo="verify", usado=False)
            .order_by("-creado_en")
            .first())

        if not cod or not cod.esta_vigente() or cod.codigo != codigo_ingresado:
            messages.error(request, "Código inválido o vencido.")
            return render(request, "registration/verificar.html")

        cod.usado = True
        cod.save()
        user.is_active = True
        user.save()

        # ahora sí: login
        login(request, user)
        del request.session["verify_user_id"]
        messages.success(request, "Cuenta verificada. Ya podés usar PozosYa.")
        return redirect("lista_baches")

    return render(request, "registration/verificar.html")



# RECUPERAR CONTRASEÑA
def _generar_codigo_6():
    return f"{secrets.randbelow(1_000_000):06d}"

def _enviar_codigo_verificacion(user, code):
    if not user.email:
        return False

    subject = "Verificá tu cuenta - PozosYa"
    message = (
        f"Hola {user.first_name or user.username},\n\n"
        f"Tu código de verificación es: {code}\n\n"
        f"Este código vence en 15 minutos.\n\n"
        f"Si no creaste esta cuenta, ignorá este correo."
    )
    send_mail(
        subject,
        message,
        settings.DEFAULT_FROM_EMAIL,
        [user.email],
        fail_silently=False,
    )
    return True



def _gen_code_6():
    return f"{secrets.randbelow(1_000_000):06d}"

def _send_reset_code_email(user, code):
    subject = "Recuperación de contraseña - PozosYa"
    message = (
        f"Hola {user.first_name or user.username},\n\n"
        f"Tu código para recuperar la contraseña es: {code}\n\n"
        f"Vence en 15 minutos.\n"
        f"Si no pediste este código, ignorá este correo."
    )
    send_mail(subject, message, settings.DEFAULT_FROM_EMAIL, [user.email], fail_silently=False)



def password_reset_request(request):
    if request.method == "POST":
        form = PasswordResetRequestForm(request.POST)
        if form.is_valid():
            email = form.cleaned_data["email"].strip().lower()

            user = User.objects.filter(email__iexact=email).first()

            # Por seguridad: siempre mostramos el mismo mensaje, exista o no.
            if user:
                code = _gen_code_6()
                PasswordResetCode.objects.create(
                    user=user,
                    code_hash=make_password(code),
                    expires_at=timezone.now() + timedelta(minutes=15),
                )
                _send_reset_code_email(user, code)

                request.session["pending_reset_user_id"] = user.id

            messages.success(request, "Si el email existe, te enviamos un código para recuperar la contraseña.")
            return redirect("password_reset_confirm")
    else:
        form = PasswordResetRequestForm()

    return render(request, "registration/password_reset_request.html", {"form": form})



def password_reset_confirm(request):
    user_id = request.session.get("pending_reset_user_id")
    if not user_id:
        messages.error(request, "No hay una recuperación pendiente. Pedí un código nuevamente.")
        return redirect("password_reset_request")

    user = User.objects.filter(id=user_id).first()
    if not user:
        messages.error(request, "No encontramos el usuario a recuperar.")
        return redirect("password_reset_request")

    if request.method == "POST":
        form = PasswordResetConfirmForm(request.POST)
        if form.is_valid():
            code = form.cleaned_data["code"].strip()
            new_password = form.cleaned_data["password1"]

            token = (
                PasswordResetCode.objects
                .filter(user=user, used_at__isnull=True)
                .order_by("-created_at")
                .first()
            )

            if not token:
                messages.error(request, "No hay un código vigente. Pedí uno nuevo.")
                return redirect("password_reset_request")

            if token.is_expired():
                messages.error(request, "El código venció. Pedí uno nuevo.")
                return redirect("password_reset_request")

            if not check_password(code, token.code_hash):
                messages.error(request, "Código incorrecto.")
                return redirect("password_reset_confirm")

            user.set_password(new_password)
            user.save()

            token.used_at = timezone.now()
            token.save(update_fields=["used_at"])

            request.session.pop("pending_reset_user_id", None)

            messages.success(request, "Contraseña actualizada. Ahora podés iniciar sesión.")
            return redirect("login")
    else:
        form = PasswordResetConfirmForm()

    return render(request, "registration/password_reset_confirm.html", {"form": form, "email": user.email})



def enviar_mail_cambio_bache(destinatario, asunto, cuerpo):
    email = (getattr(destinatario, "email", "") or "").strip()
    if not email:
        return  # no hay email

    # Si tenés verificación en Perfil y querés exigirla:
    # (cambiá "email_verificado" por el nombre real de tu campo si difiere)
    if hasattr(destinatario, "perfil") and hasattr(destinatario.perfil, "email_verificado"):
        if not destinatario.perfil.email_verificado:
            return

    send_mail(
        subject=asunto,
        message=cuerpo,
        from_email=getattr(settings, "DEFAULT_FROM_EMAIL", None),
        recipient_list=[email],
        fail_silently=True,  # si querés ver errores, ponelo en False mientras probás
    )



